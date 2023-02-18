#!/usr/bin/env python3
import warnings
import logging
import sys
import os
from datetime import datetime

import openpyxl
from dotenv import dotenv_values

import calendar_draw as cd

# setup logging
FORMATTER = logging.Formatter("%(asctime)s — %(name)s — %(levelname)s — %(message)s")
log = logging.getLogger(__name__)
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(FORMATTER)
# file_handler = logging.FileHandler(LOG_FILE)
# file_handler.setFormatter(FORMATTER)
log.addHandler(console_handler)
# log.addHandler(file_handler)
if "DEBUG" in os.environ:
    log.setLevel(logging.DEBUG)
else:
    log.setLevel(logging.INFO)


def main():
    # Load users from .env file
    config = dotenv_values(".env")

    teams, countries = cd.build_dict(config)
    log.debug(f"countries: {countries}")
    log.debug(f"teams: {teams}")

    cd.draw(datetime.now(), teams)
    # Define variable to load the dataframe
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        dataframe = openpyxl.load_workbook("AMZ-SREG.xlsx")

    # Define variable to read sheet
    dataframe1 = dataframe.active

    # Iterate the loop to read the cell values
    for row in list(range(0, dataframe1.max_row + 1))[0:5]:
        for col in list(dataframe1.iter_cols(1, dataframe1.max_column + 1))[-6:-1]:
            print(col[row].value, end=":")
        print(" ")


if __name__ == "__main__":
    main()
